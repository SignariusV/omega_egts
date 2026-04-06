"""
Показать порядок листов в файле 1 и первые 6 пакетов
"""
import openpyxl
from pathlib import Path

DOCS_DIR = Path(__file__).parent.parent / "docs"
file_path = DOCS_DIR / "1. Состав команд для конфигурирования (OK).xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)

print(f"📁 {file_path.name}")
print(f"📋 Всего листов: {len(wb.sheetnames)}\n")

print("Порядок листов:")
for i, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    
    # Считаем байты пакета
    byte_count = 0
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None and str(val).strip():
            byte_count += 1
    
    direction = "PLATFORM → УСВ" if any(k in name.upper() for k in ['RESPONSE', 'COMCONF', 'SERVER', 'APN']) else "УСВ → PLATFORM"
    
    print(f"  {i:2d}. {name:<45} ({byte_count:3d} байт) {direction}")

wb.close()
