"""
Показать процедуру аутентификации с направлениями и hex данными
"""
import openpyxl
from pathlib import Path

DOCS_DIR = Path(__file__).parent.parent / "docs"
file_path = DOCS_DIR / "2. Передача траектории движения (ОК).xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)

# Листы 3-8: аутентификация (TERM_IDENTITY + RESPONSE + VEHICLE_DATA + RESPONSE + RESULT_CODE + RESPONSE)
sheets = wb.sheetnames[2:8]

print("="*80)
print("🔐 ПРОЦЕДУРА АУТЕНТИФИКАЦИИ (TCP/IP)")
print("Файл: 2. Передача траектории движения (ОК).xlsx")
print("="*80)

for i, sheet_name in enumerate(sheets, 1):
    ws = wb[sheet_name]
    
    # Собираем hex и описание
    hex_bytes = []
    fields = []
    
    for row in range(1, ws.max_row + 1):
        hex_val = ws.cell(row=row, column=1).value
        field_name = ws.cell(row=row, column=2).value
        field_type = ws.cell(row=row, column=3).value
        
        if hex_val is not None and str(hex_val).strip():
            hex_bytes.append(str(hex_val).strip().upper())
        
        if field_name:
            parts = [str(field_name)]
            if field_type:
                parts.append(f"[{field_type}]")
            fields.append(" ".join(parts))
    
    full_hex = "".join(hex_bytes)
    size = len(hex_bytes)
    
    # Направление
    if "RESPONSE" in sheet_name.upper() or "RESULT_CODE" in sheet_name.upper():
        direction = "ПЛАТФОРМА → УСВ"
    elif "TERM_IDENTITY" in sheet_name.upper() or "VEHICLE_DATA" in sheet_name.upper():
        direction = "УСВ → ПЛАТФОРМА"
    else:
        direction = "Не определено"
    
    print(f"\n{'─'*80}")
    print(f"Шаг {i}: {sheet_name}")
    print(f"{'─'*80}")
    print(f"  Направление: {direction}")
    print(f"  Размер: {size} байт ({len(full_hex)} hex символов)")
    print(f"  HEX: {full_hex[:120]}...")
    print(f"\n  Поля:")
    for f in fields[:8]:
        print(f"    • {f}")
    if len(fields) > 8:
        print(f"    ... и ещё {len(fields) - 8}")

wb.close()
