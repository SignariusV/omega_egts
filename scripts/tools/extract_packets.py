"""
Скрипт для извлечения всех hex пакетов из xlsx файлов docs/
Извлекает:
  - Чистые hex данные
  - Подробное описание каждого пакета
  - Статистику по пакетам
"""
import os
import json
import openpyxl
from pathlib import Path
from collections import defaultdict, Counter
from datetime import datetime

DOCS_DIR = Path(__file__).parent.parent / "docs"
OUTPUT_DIR = Path(__file__).parent.parent / "data" / "packets"

def extract_packets():
    """Извлекает все пакеты из xlsx файлов"""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    xlsx_files = sorted(DOCS_DIR.glob("*.xlsx"))
    
    all_packets = []  # Все пакеты с описаниями
    pure_hex_list = []  # Только чистые hex
    packet_stats = {
        'by_type': Counter(),
        'by_file': Counter(),
        'by_direction': Counter(),
        'total_packets': 0,
        'unique_hex': set(),
    }
    
    print(f"🔍 Извлечение пакетов из {len(xlsx_files)} файлов...\n")
    
    for file_idx, file_path in enumerate(xlsx_files, 1):
        print(f"[{file_idx}/{len(xlsx_files)}] 📁 {file_path.name}")
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
                ws = wb[sheet_name]
                max_row = ws.max_row
                max_col = ws.max_column
                
                print(f"  📄 Лист {sheet_idx}/{len(wb.sheetnames)}: {sheet_name} ({max_row} строк)")
                
                # Определяем направление по имени листа
                direction = detect_direction(sheet_name)
                packet_stats['by_direction'][direction] += 1
                
                # Читаем все строки
                packet_count = 0
                for row_idx in range(1, max_row + 1):
                    hex_value = ws.cell(row=row_idx, column=1).value
                    
                    if not hex_value:
                        continue
                    
                    hex_value = str(hex_value).strip()
                    
                    # Собираем описание из остальных колонок
                    description_parts = []
                    for col_idx in range(2, max_col + 1):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if cell_value is not None:
                            description_parts.append(str(cell_value))
                    
                    description = " | ".join(description_parts) if description_parts else ""
                    
                    # Создаём запись пакета
                    packet = {
                        'file': file_path.name,
                        'sheet': sheet_name,
                        'row': row_idx,
                        'hex': hex_value,
                        'description': description,
                        'direction': direction,
                        'hex_length': len(hex_value.replace(' ', '').replace(':', '')) // 2,
                    }
                    
                    all_packets.append(packet)
                    pure_hex_list.append(hex_value)
                    
                    # Статистика
                    packet_type = sheet_name.split('(')[0].strip()
                    packet_stats['by_type'][packet_type] += 1
                    packet_stats['by_file'][file_path.name] += 1
                    packet_stats['total_packets'] += 1
                    
                    # Нормализуем hex для подсчёта уникальных
                    normalized_hex = hex_value.replace(' ', '').replace(':', '').upper()
                    packet_stats['unique_hex'].add(normalized_hex)
                    
                    packet_count += 1
                
                print(f"     ✅ Извлечено пакетов: {packet_count}")
            
            wb.close()
            print()
            
        except Exception as e:
            print(f"  ❌ Ошибка чтения файла: {e}\n")
    
    return all_packets, pure_hex_list, packet_stats


def detect_direction(sheet_name):
    """Определяет направление передачи по имени листа"""
    sheet_upper = sheet_name.upper()
    
    # Сервер -> Устройство (ответы, подтверждения)
    server_keywords = ['RESPONSE', 'RESULT_CODE', 'COMCONF', 'SERVER', 'APN', 'ADDRESS']
    # Устройство -> Сервер (данные с устройства)
    device_keywords = ['TRACK_DATA', 'ACCEL_DATA', 'VEHICLE_DATA', 'SERVICE_PART', 'COMMAND_DATA']
    
    for keyword in server_keywords:
        if keyword in sheet_upper:
            return 'PLATFORM -> УСВ'
    
    for keyword in device_keywords:
        if keyword in sheet_upper:
            return 'УСВ -> PLATFORM'
    
    return 'Не определено'


def save_results(all_packets, pure_hex_list, packet_stats):
    """Сохраняет результаты в файлы"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 1. Все пакеты с описаниями (JSON)
    packets_file = OUTPUT_DIR / f"all_packets_{timestamp}.json"
    with open(packets_file, 'w', encoding='utf-8') as f:
        json.dump(all_packets, f, ensure_ascii=False, indent=2)
    print(f"\n💾 Сохранено: {packets_file}")
    print(f"   Всего пакетов: {len(all_packets)}")
    
    # 2. Чистые hex (TXT)
    hex_file = OUTPUT_DIR / f"pure_hex_{timestamp}.txt"
    with open(hex_file, 'w', encoding='utf-8') as f:
        for hex_val in pure_hex_list:
            f.write(hex_val + '\n')
    print(f"\n💾 Сохранено: {hex_file}")
    print(f"   Всего HEX записей: {len(pure_hex_list)}")
    
    # 3. Статистика (JSON)
    stats_file = OUTPUT_DIR / f"statistics_{timestamp}.json"
    stats_serializable = {
        'total_packets': packet_stats['total_packets'],
        'unique_hex_count': len(packet_stats['unique_hex']),
        'by_type': dict(packet_stats['by_type']),
        'by_file': dict(packet_stats['by_file']),
        'by_direction': dict(packet_stats['by_direction']),
        'top_20_types': packet_stats['by_type'].most_common(20),
    }
    with open(stats_file, 'w', encoding='utf-8') as f:
        json.dump(stats_serializable, f, ensure_ascii=False, indent=2)
    print(f"\n💾 Сохранено: {stats_file}")


def print_statistics(packet_stats):
    """Выводит статистику в консоль"""
    print("\n" + "="*80)
    print("📊 СТАТИСТИКА ПАКЕТОВ")
    print("="*80)
    
    print(f"\n📦 Всего пакетов: {packet_stats['total_packets']}")
    print(f"🔢 Уникальных HEX: {len(packet_stats['unique_hex'])}")
    
    print(f"\n🔄 По направлению передачи:")
    for direction, count in packet_stats['by_direction'].items():
        print(f"   {direction}: {count}")
    
    print(f"\n📋 По типам пакетов (ТОП-20):")
    for packet_type, count in packet_stats['by_type'].most_common(20):
        print(f"   {packet_type}: {count}")
    
    print(f"\n📁 По файлам:")
    for file_name, count in packet_stats['by_file'].items():
        print(f"   {file_name}: {count}")


if __name__ == "__main__":
    all_packets, pure_hex_list, packet_stats = extract_packets()
    
    print_statistics(packet_stats)
    save_results(all_packets, pure_hex_list, packet_stats)
    
    print("\n✅ Извлечение завершено!")
