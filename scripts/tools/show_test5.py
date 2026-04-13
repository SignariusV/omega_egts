"""
Показать структуру файла 5
"""
import openpyxl
from pathlib import Path

DOCS_DIR = Path(__file__).parent.parent / "docs"
file_path = DOCS_DIR / "5. Передача настроек для УСВ (OK).xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)

print(f"📁 {file_path.name}")
print(f"📋 Всего листов: {len(wb.sheetnames)}\n")

print("Все листы:")
for i, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    
    # Считаем байты
    byte_count = 0
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None and str(val).strip():
            byte_count += 1
    
    print(f"  {i:2d}. {name:<45} ({byte_count:3d} байт)")

wb.close()

print(f"\n{'='*80}")
print("📊 СТРУКТУРА ФАЙЛА 5:")
print(f"{'='*80}")
print("""
Листы 1-6: Аутентификация (аналогична файлам 1-4)
  1. EGTS_SR_TERM_IDENTITY (1)     — Идентификация терминала
  2. EGTS_SR_RECORD_RESPONSE (2)   — Подтверждение
  3. EGTS_SR_VEHICLE_DATA (3)      — Данные ТС
  4. EGTS_SR_RECORD_RESPONSE (4)   — Подтверждение
  5. EGTS_SR_RESULT_CODE (5)       — Результат
  6. EGTS_SR_RECORD_RESPONSE (6)   — Подтверждение

Лист 7: Основной тест — команды конфигурирования
  7. EGTS_SR_COMMAND_DATA (7)      — Данные команды (72 байт)

Лист 8: Транспортный ответ
  8. EGTS_PT_RESPONSE (8)          — Ответ транспортного уровня (29 байт)

Лист 9: Подтверждение команды
  9. CT_COMCONF (9)                — Подтверждение выполнения команды (49 байт)
""")
