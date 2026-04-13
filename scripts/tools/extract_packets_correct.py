"""
Правильное извлечение пакетов из xlsx
Каждый лист = один или несколько пакетов
Колонка A = hex байта, остальные колонки = описание
"""
import openpyxl
import json
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime

DOCS_DIR = Path(__file__).parent.parent / "docs"
OUTPUT_DIR = Path(__file__).parent.parent / "data" / "packets"

def extract_packets_correct():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    xlsx_files = sorted(DOCS_DIR.glob("*.xlsx"))
    
    all_packets = []
    packet_counter = Counter()
    direction_counter = Counter()
    file_counter = Counter()
    
    print(f"🔍 Извлечение пакетов из {len(xlsx_files)} файлов...\n")
    
    for file_idx, file_path in enumerate(xlsx_files, 1):
        print(f"[{file_idx}/{len(xlsx_files)}] 📁 {file_path.name}")
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
                ws = wb[sheet_name]
                
                # Находим границы пакетов (по PRV)
                packet_boundaries = find_packet_boundaries(ws)
                
                print(f"  📄 {sheet_name}: {len(packet_boundaries)} пакет(ов)")
                
                for pkt_num, (start_row, end_row) in enumerate(packet_boundaries, 1):
                    # Собираем полный hex пакета
                    hex_bytes = []
                    description_parts = []
                    
                    for row_idx in range(start_row, end_row + 1):
                        hex_val = ws.cell(row=row_idx, column=1).value
                        field_name = ws.cell(row=row_idx, column=2).value
                        field_type = ws.cell(row=row_idx, column=3).value
                        field_desc = ws.cell(row=row_idx, column=4).value
                        
                        if hex_val is not None:
                            hex_bytes.append(str(hex_val).strip().upper())
                        
                        # Собираем описание поля
                        if field_name:
                            desc_parts = [str(field_name)]
                            if field_type:
                                desc_parts.append(f"[{field_type}]")
                            if field_desc:
                                desc_parts.append(str(field_desc))
                            description_parts.append(" ".join(desc_parts))
                    
                    full_hex = "".join(hex_bytes)
                    full_description = "\n".join(description_parts)
                    
                    # Определяем направление
                    direction = detect_direction(sheet_name)
                    
                    packet = {
                        'file': file_path.name,
                        'sheet': sheet_name,
                        'packet_num_in_sheet': pkt_num,
                        'rows': f"{start_row}-{end_row}",
                        'hex': full_hex,
                        'hex_length_bytes': len(hex_bytes),
                        'description': full_description,
                        'direction': direction,
                    }
                    
                    all_packets.append(packet)
                    packet_counter[sheet_name] += 1
                    direction_counter[direction] += 1
                    file_counter[file_path.name] += 1
                
                print(f"     ✅ Извлечено пакетов: {len(packet_boundaries)}")
            
            wb.close()
            print()
            
        except Exception as e:
            print(f"  ❌ Ошибка: {e}\n")
    
    return all_packets, packet_counter, direction_counter, file_counter


def find_packet_boundaries(ws):
    """Находит границы пакетов по PRV (Protocol Version)"""
    prv_rows = []
    
    for row_idx in range(1, ws.max_row + 1):
        hex_val = ws.cell(row=row_idx, column=1).value
        desc_val = ws.cell(row=row_idx, column=2).value
        
        if hex_val is not None and str(hex_val).strip().upper() == '01':
            if desc_val and 'PRV' in str(desc_val).upper():
                prv_rows.append(row_idx)
    
    if not prv_rows:
        # Если нет PRV, считаем весь лист одним пакетом
        # Ищем первую и последнюю строку с данными
        first = 1
        last = ws.max_row
        while last > first:
            has_data = any(ws.cell(row=last, column=c).value for c in range(1, min(ws.max_column+1, 10)))
            if has_data:
                break
            last -= 1
        return [(first, last)]
    
    boundaries = []
    for i, start in enumerate(prv_rows):
        end = prv_rows[i+1] - 1 if i+1 < len(prv_rows) else ws.max_row
        # Убираем пустые строки в конце
        while end > start:
            has_data = any(ws.cell(row=end, column=c).value for c in range(1, min(ws.max_column+1, 10)))
            if has_data:
                break
            end -= 1
        boundaries.append((start, end))
    
    return boundaries


def detect_direction(sheet_name):
    """Определяет направление"""
    sheet_upper = sheet_name.upper()
    
    server_to_device = ['RESPONSE', 'RESULT_CODE', 'COMCONF', 'SERVER', 'APN']
    device_to_server = ['TRACK_DATA', 'ACCEL_DATA', 'VEHICLE_DATA', 'SERVICE_PART', 'COMMAND_DATA']
    
    for kw in server_to_device:
        if kw in sheet_upper:
            return 'PLATFORM -> УСВ'
    for kw in device_to_server:
        if kw in sheet_upper:
            return 'УСВ -> PLATFORM'
    
    return 'Не определено'


def save_and_report(all_packets, packet_counter, direction_counter, file_counter):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 1. Все пакеты
    packets_file = OUTPUT_DIR / f"all_packets_correct_{timestamp}.json"
    with open(packets_file, 'w', encoding='utf-8') as f:
        json.dump(all_packets, f, ensure_ascii=False, indent=2)
    
    # 2. Только hex
    hex_file = OUTPUT_DIR / f"pure_hex_correct_{timestamp}.txt"
    with open(hex_file, 'w', encoding='utf-8') as f:
        for pkt in all_packets:
            f.write(pkt['hex'] + '\n')
    
    # 3. Статистика
    stats = {
        'total_packets': len(all_packets),
        'unique_hex_count': len(set(p['hex'] for p in all_packets)),
        'by_type': dict(packet_counter),
        'by_direction': dict(direction_counter),
        'by_file': dict(file_counter),
    }
    stats_file = OUTPUT_DIR / f"statistics_correct_{timestamp}.json"
    with open(stats_file, 'w', encoding='utf-8') as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    
    # Вывод статистики
    print("\n" + "="*80)
    print("📊 СТАТИСТИКА (ПРАВИЛЬНАЯ)")
    print("="*80)
    
    print(f"\n📦 Всего пакетов: {len(all_packets)}")
    print(f"🔢 Уникальных HEX: {stats['unique_hex_count']}")
    
    print(f"\n🔄 По направлению:")
    for dir, count in direction_counter.most_common():
        print(f"   {dir}: {count}")
    
    print(f"\n📋 По типам пакетов:")
    for type, count in packet_counter.most_common():
        print(f"   {type}: {count}")
    
    print(f"\n📁 По файлам:")
    for file, count in file_counter.most_common():
        print(f"   {file}: {count}")
    
    # Размеры пакетов
    print(f"\n📏 Размеры пакетов:")
    sizes = [p['hex_length_bytes'] for p in all_packets]
    print(f"   Мин: {min(sizes)} байт")
    print(f"   Макс: {max(sizes)} байт")
    print(f"   Средний: {sum(sizes) // len(sizes)} байт")
    
    print(f"\n💾 Сохранено:")
    print(f"   {packets_file}")
    print(f"   {hex_file}")
    print(f"   {stats_file}")


if __name__ == "__main__":
    all_packets, packet_counter, direction_counter, file_counter = extract_packets_correct()
    save_and_report(all_packets, packet_counter, direction_counter, file_counter)
    print("\n✅ Завершено!")
