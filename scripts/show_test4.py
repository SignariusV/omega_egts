"""
Показать структуру файла 4
"""
import openpyxl
from pathlib import Path

DOCS_DIR = Path(__file__).parent.parent / "docs"
file_path = DOCS_DIR / "4. Передача обновлённого ПО (OK).xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)

print(f"📁 {file_path.name}")
print(f"📋 Всего листов: {len(wb.sheetnames)}\n")

print("Все листы:")
for i, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    
    # Считаем байты
    byte_count = 0
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None and str(val).strip():
            byte_count += 1
    
    print(f"  {i:2d}. {name:<45} ({byte_count:3d} байт)")

wb.close()

print(f"\n{'='*80}")
print("📊 СТРУКТУРА ФАЙЛА 4:")
print(f"{'='*80}")
print("""
Листы 1-6: Аутентификация (аналогична файлам 1-3)
  1. EGTS_SR_TERM_IDENTITY (1)     — Идентификация терминала
  2. EGTS_SR_RECORD_RESPONSE (2)   — Подтверждение
  3. EGTS_SR_VEHICLE_DATA (3)      — Данные ТС
  4. EGTS_SR_RECORD_RESPONSE (4)   — Подтверждение
  5. EGTS_SR_RESULT_CODE (5)       — Результат
  6. EGTS_SR_RECORD_RESPONSE (6)   — Подтверждение

Листы 7-8: Передача первой части ПО
  7. EGTS_SR_SERVICE_PART_DATA (7) — Данные ПО (51 236 байт!)
  8. EGTS_SR_RECORD_RESPONSE (8)   — Подтверждение

Листы 9-10: Передача второй части ПО
  9. EGTS_SR_SERVICE_PART_DATA (9) — Данные ПО (29 312 байт!)
 10. EGTS_SR_RECORD_RESPONSE (10)  — Подтверждение
""")
