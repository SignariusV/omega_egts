"""
Скрипт для анализа листов во всех xlsx файлах docs/
Показывает: имя листа, количество строк, количество колонок
"""
import os
import openpyxl
from pathlib import Path

DOCS_DIR = Path(__file__).parent.parent / "docs"

def analyze_xlsx_files():
    xlsx_files = list(DOCS_DIR.glob("*.xlsx"))
    
    if not xlsx_files:
        print("❌ XLSX файлы не найдены")
        return
    
    print(f"📊 Найдено XLSX файлов: {len(xlsx_files)}\n")
    
    for file_path in xlsx_files:
        print(f"\n{'='*80}")
        print(f"📁 Файл: {file_path.name}")
        print(f"{'='*80}")
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Определяем размеры
                max_row = ws.max_row
                max_col = ws.max_column
                
                print(f"\n  📄 Лист: '{sheet_name}'")
                print(f"     Строки: {max_row}, Колонки: {max_col}")
                
                # Пример первых 3 строк из колонки A (hex)
                print(f"     Примеры HEX (колонка A):")
                for row_idx in range(1, min(4, max_row + 1)):
                    cell_value = ws.cell(row=row_idx, column=1).value
                    if cell_value:
                        print(f"       Строка {row_idx}: {cell_value[:80]}..." if len(str(cell_value)) > 80 else f"       Строка {row_idx}: {cell_value}")
            
            wb.close()
            
        except Exception as e:
            print(f"  ❌ Ошибка чтения файла: {e}")

if __name__ == "__main__":
    analyze_xlsx_files()
