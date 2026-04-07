"""
Показать что осталось в файле 1 после исключения верификации (листы 1-6)
"""
import openpyxl
from pathlib import Path

DOCS_DIR = Path(__file__).parent.parent / "docs"
file_path = DOCS_DIR / "1. Состав команд для конфигурирования (OK).xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)

print(f"📁 {file_path.name}")
print(f"📋 Всего листов: {len(wb.sheetnames)}\n")

print("Все листы:")
for i, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    
    # Считаем байты
    byte_count = 0
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None and str(val).strip():
            byte_count += 1
    
    print(f"  {i:2d}. {name:<45} ({byte_count:3d} байт)")

wb.close()

print(f"\n{'='*80}")
print("📊 СТРУКТУРА ФАЙЛА 1:")
print(f"{'='*80}")
print(f"""
Листы 1-6: Процедура верификации (SMS)
  1. EGTS_GPRS_APN (SMS)(1)        — Настройки APN
  2. CT_COMCONF (SMS)(2)           — Подтверждение
  3. EGTS_SERVER_ADDRESS (SMS)(3)  — Адрес сервера
  4. CT_COMCONF (SMS)(4)           — Подтверждение
  5. EGTS_UNIT_ID (SMS)(5)         — Запрос UNIT_ID
  6. CT_COMCONF (SMS)(6)           — Ответ с UNIT_ID

Листы 7-12: ???
  7. EGTS_SR_TERM_IDENTITY (7)     — Идентификация терминала
  8. EGTS_SR_RECORD_RESPONSE (8)   — Подтверждение
  9. EGTS_SR_VEHICLE_DATA (9)      — Данные ТС
 10. EGTS_SR_RECORD_RESPONSE (10)  — Подтверждение
 11. EGTS_SR_RESULT_CODE (11)      — Результат
 12. EGTS_SR_RECORD_RESPONSE (12)  — Подтверждение
""")
