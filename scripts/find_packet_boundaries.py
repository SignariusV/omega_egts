"""
Определить: сколько пакетов в маленьких листах?
Ищем границы пакетов — повторяющиеся структуры (заголовки, PRV)
"""
import openpyxl
from pathlib import Path

DOCS_DIR = Path(__file__).parent.parent / "docs"

def find_packet_boundaries():
    file_path = DOCS_DIR / "2. Передача траектории движения (ОК).xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Маленький лист
    sheet_name = "EGTS_TRACK_DATA (1)"
    ws = wb[sheet_name]
    
    print(f"📄 {sheet_name} ({ws.max_row} строк × {ws.max_column} колонок)\n")
    
    # Ищем строки где колонка 1 = '01' (PRV - начало пакета)
    print("🔍 Строки где колонка A (hex) = '01' (PRV - начало пакета):")
    prv_rows = []
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val is not None and str(val).strip() == '01':
            prv_rows.append(row_idx)
            # Показываем описание
            desc = ws.cell(row=row_idx, column=2).value
            print(f"  Строка {row_idx}: {desc}")
    
    print(f"\n  Найдено {len(prv_rows)} начал пакетов")
    print(f"  Границы пакетов: строки {prv_rows}")
    
    # Определяем пакеты
    print(f"\n📦 ГРАНИЦЫ ПАКЕТОВ:")
    for i, start in enumerate(prv_rows):
        end = prv_rows[i+1] - 1 if i+1 < len(prv_rows) else ws.max_row
        size = end - start + 1
        print(f"  Пакет {i+1}: строки {start}-{end} ({size} байт)")
    
    # Теперь проверим все листы в файле 1
    print("\n\n" + "="*80)
    print("📋 АНАЛИЗ ВСЕХ ЛИСТОВ: поиск пакетов")
    print("="*80)
    
    file_path1 = DOCS_DIR / "1. Состав команд для конфигурирования (OK).xlsx"
    wb1 = openpyxl.load_workbook(file_path1, data_only=True)
    
    for sheet_name in wb1.sheetnames:
        ws = wb1[sheet_name]
        
        # Считаем сколько раз встречается '01' в колонке A с описанием PRV
        prv_count = 0
        for row_idx in range(1, ws.max_row + 1):
            hex_val = ws.cell(row=row_idx, column=1).value
            desc_val = ws.cell(row=row_idx, column=2).value
            if hex_val is not None and str(hex_val).strip() == '01':
                if desc_val and 'PRV' in str(desc_val).upper():
                    prv_count += 1
        
        print(f"\n  📄 {sheet_name} ({ws.max_row} строк)")
        print(f"     Найдено PRV (начал пакетов): {prv_count}")
        
        # Если пакетов несколько, покажем границы
        if prv_count > 1:
            prv_rows = []
            for row_idx in range(1, ws.max_row + 1):
                hex_val = ws.cell(row=row_idx, column=1).value
                desc_val = ws.cell(row=row_idx, column=2).value
                if hex_val is not None and str(hex_val).strip() == '01' and desc_val and 'PRV' in str(desc_val).upper():
                    prv_rows.append(row_idx)
            
            print(f"     Границы: строки {prv_rows}")
            for i, start in enumerate(prv_rows):
                # Ищем конец пакета (перед следующим PRV или конец листа)
                end = prv_rows[i+1] - 1 if i+1 < len(prv_rows) else ws.max_row
                # Ищем последнюю строку с данными
                while end > start:
                    has_data = any(ws.cell(row=end, column=c).value for c in range(1, ws.max_column + 1))
                    if has_data:
                        break
                    end -= 1
                size = end - start + 1
                print(f"       Пакет {i+1}: строки {start}-{end} ({size} байт)")
    
    wb.close()
    wb1.close()


if __name__ == "__main__":
    find_packet_boundaries()
