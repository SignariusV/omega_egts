"""
Понять структуру: один пакет = все строки листа или каждый лист = набор пакетов?
Смотрим на EGTS_SR_TRACK_DATA (9) - 883 строки, 880 колонок
"""
import openpyxl
from pathlib import Path

DOCS_DIR = Path(__file__).parent.parent / "docs"

def inspect_big_sheet():
    file_path = DOCS_DIR / "2. Передача траектории движения (ОК).xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    sheet_name = "EGTS_SR_TRACK_DATA (9)"
    ws = wb[sheet_name]
    
    print(f"Лист: {sheet_name}")
    print(f"Строк: {ws.max_row}, Колонн: {ws.max_column}\n")
    
    # Проверим: сколько колонок заполнено в разных строках
    print("📊 Заполненность колонок (первые 10 строк):")
    for row_idx in range(1, 11):
        filled = 0
        first_vals = []
        for col in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(row=row_idx, column=col).value
            if val is not None and str(val).strip():
                filled += 1
                first_vals.append(str(val)[:40])
        
        print(f"  Строка {row_idx}: заполнено {filled} колонок | {first_vals[:5]}")
    
    # Проверим последние строки
    print(f"\n📊 Заполненность колонок (строки 878-883):")
    for row_idx in range(max(1, ws.max_row - 5), ws.max_row + 1):
        filled = 0
        first_vals = []
        for col in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(row=row_idx, column=col).value
            if val is not None and str(val).strip():
                filled += 1
                first_vals.append(str(val)[:40])
        
        print(f"  Строка {row_idx}: заполнено {filled} колонок | {first_vals[:5]}")
    
    # Ключевой вопрос: в каких колонках есть данные?
    # Проверим строку 1 - где данные?
    print(f"\n📋 Полная строка 1 (первые 30 колонок):")
    for col in range(1, 31):
        val = ws.cell(row=1, column=col).value
        if val is not None:
            print(f"  [{col:3d}] {str(val)[:60]}")
    
    # Проверим строку где много колонок
    print(f"\n📋 Строка 500 — проверяем заполненность колонок:")
    filled_cols = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=500, column=col).value
        if val is not None and str(val).strip():
            filled_cols.append(col)
    
    print(f"  Заполненные колонки: {filled_cols[:20]}{'...' if len(filled_cols) > 20 else ''}")
    print(f"  Всего заполнено: {len(filled_cols)} из {ws.max_column}")
    
    # Проверяем: есть ли строки с большим количеством колонок?
    print(f"\n📊 Максимальная заполненность по строкам:")
    max_filled = 0
    max_row = 0
    for row_idx in range(1, ws.max_row + 1):
        filled = 0
        for col in range(1, ws.max_column + 1, 50):  # шаг 50 для скорости
            val = ws.cell(row=row_idx, column=col).value
            if val is not None and str(val).strip():
                filled += 1
        if filled > max_filled:
            max_filled = filled
            max_row = row_idx
    
    print(f"  Строка {max_row} имеет ~{max_filled} заполненных колонок (с шагом 50)")
    
    wb.close()


def compare_sheets():
    """Сравним маленький и большой лист"""
    file_path = DOCS_DIR / "2. Передача траектории движения (ОК).xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    print("\n\n" + "="*80)
    print("📋 СРАВНЕНИЕ ЛИСТОВ")
    print("="*80)
    
    for sheet_name in wb.sheetnames[:3]:
        ws = wb[sheet_name]
        print(f"\n📄 {sheet_name} ({ws.max_row} строк × {ws.max_column} колонок)")
        
        # Сколько строк имеют > 1 заполненной колонки
        rows_with_data = 0
        for row_idx in range(1, ws.max_row + 1):
            filled = sum(1 for col in range(1, min(ws.max_column + 1, 30)) 
                        if ws.cell(row=row_idx, column=col).value is not None)
            if filled > 3:
                rows_with_data += 1
        
        print(f"   Строк с данными (>3 колонок): {rows_with_data}")
    
    wb.close()


if __name__ == "__main__":
    inspect_big_sheet()
    compare_sheets()
