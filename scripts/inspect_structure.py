"""
Скрипт для изучения структуры данных в xlsx
Покажем: что в колонках для нескольких строк
"""
import openpyxl
from pathlib import Path

DOCS_DIR = Path(__file__).parent.parent / "docs"

def inspect_structure():
    # Берём маленький файл для начала
    file_path = DOCS_DIR / "1. Состав команд для конфигурирования (OK).xlsx"
    
    print(f"📁 {file_path.name}\n")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Берём первый лист
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    
    print(f"Лист: {sheet_name}")
    print(f"Строк: {ws.max_row}, Колонн: {ws.max_column}\n")
    
    # Заголовки (первая строка)
    print("📋 ЗАГОЛОВКИ (строка 1):")
    headers = []
    for col in range(1, min(ws.max_column + 1, 30)):
        val = ws.cell(row=1, column=col).value
        headers.append(val)
        print(f"  Колонка {col}: {val}")
    
    # Данные: первые 5 строк данных
    print(f"\n📦 ДАННЫЕ (строки 2-6):")
    for row_idx in range(2, 7):
        print(f"\n  --- Строка {row_idx} ---")
        row_hex_parts = []
        for col in range(1, min(ws.max_column + 1, 30)):
            val = ws.cell(row=row_idx, column=col).value
            if val is not None:
                val_str = str(val).strip()
                if val_str:
                    row_hex_parts.append(val_str)
                    # Показываем первые 15 колонок
                    if col <= 15:
                        print(f"  Колонка {col:2d}: {val_str[:50]}")
        
        # Собираем полный hex
        full_hex = " ".join(row_hex_parts)
        print(f"\n  ✅ ПОЛНЫЙ HEX ({len(row_hex_parts)} частей):")
        print(f"     {full_hex[:200]}{'...' if len(full_hex) > 200 else ''}")
    
    wb.close()
    
    # Теперь посмотрим на большой лист
    print("\n\n" + "="*80)
    file_path2 = DOCS_DIR / "2. Передача траектории движения (ОК).xlsx"
    print(f"📁 {file_path2.name}")
    
    wb2 = openpyxl.load_workbook(file_path2, data_only=True)
    
    # Лист с EGTS_SR_TRACK_DATA (9) - большой
    sheet_name2 = "EGTS_SR_TRACK_DATA (9)"
    if sheet_name2 in wb2.sheetnames:
        ws2 = wb2[sheet_name2]
        print(f"\nЛист: {sheet_name2}")
        print(f"Строк: {ws2.max_row}, Колонн: {ws2.max_column}\n")
        
        # Заголовки
        print("📋 ЗАГОЛОВКИ (строка 1):")
        for col in range(1, min(ws2.max_column + 1, 15)):
            val = ws2.cell(row=1, column=col).value
            if val:
                print(f"  Колонка {col}: {val}")
        
        # Первые 3 строки данных
        print(f"\n📦 ДАННЫЕ (строки 2-4):")
        for row_idx in range(2, 5):
            print(f"\n  --- Строка {row_idx} ---")
            for col in range(1, min(ws2.max_column + 1, 15)):
                val = ws2.cell(row=row_idx, column=col).value
                if val is not None:
                    val_str = str(val).strip()
                    if val_str:
                        print(f"  Колонка {col:2d}: {val_str[:80]}")
    
    wb2.close()


if __name__ == "__main__":
    inspect_structure()
