"""
Показать структуру файла 2
"""
import openpyxl
from pathlib import Path

DOCS_DIR = Path(__file__).parent.parent / "docs"
file_path = DOCS_DIR / "2. Передача траектории движения (ОК).xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)

print(f"📁 {file_path.name}")
print(f"📋 Всего листов: {len(wb.sheetnames)}\n")

print("Все листы:")
for i, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    
    # Считаем байты
    byte_count = 0
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None and str(val).strip():
            byte_count += 1
    
    print(f"  {i:2d}. {name:<45} ({byte_count:3d} байт)")

wb.close()

print(f"\n{'='*80}")
print("📊 СТРУКТУРА ФАЙЛА 2:")
print(f"{'='*80}")
print("""
Листы 1-2: Передача траектории
  1. EGTS_TRACK_DATA (1)           — Запрос/данные траектории?
  2. CT_COMCONF (2)                — Подтверждение

Листы 3-8: Аутентификация (аналогична файлу 1)
  3. EGTS_SR_TERM_IDENTITY (3)     — Идентификация терминала
  4. EGTS_SR_RECORD_RESPONSE (4)   — Подтверждение
  5. EGTS_SR_VEHICLE_DATA (5)      — Данные ТС
  6. EGTS_SR_RECORD_RESPONSE (6)   — Подтверждение
  7. EGTS_SR_RESULT_CODE (7)       — Результат
  8. EGTS_SR_RECORD_RESPONSE (8)   — Подтверждение

Лист 9: Основной тест
  9. EGTS_SR_TRACK_DATA (9)        — Траектория движения (880 байт!)

Лист 10: Завершение
 10. EGTS_SR_RECORD_RESPONSE (10)  — Подтверждение
""")
